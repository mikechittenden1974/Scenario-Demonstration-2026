import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from datetime import date
import pickle, os

OUT = "/sessions/tender-sleepy-hawking/mnt/AM1_2/CosmoCourier_SD2_Training_Pilot"

# Load records
with open(f"{OUT}/records.pkl","rb") as f:
    records = pickle.load(f)

# ── Style helpers ────────────────────────────────────────────────────────────
DARK  = "1F3864"
BLUE  = "2F5496"
LBLUE = "D5E8F0"
ACCENT= "F4B942"
GREEN = "E2EFDA"
LGRAY = "F2F2F2"
WHITE = "FFFFFF"
RED   = "FF0000"

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center")

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ── Write the Deliveries data sheet ─────────────────────────────────────────
HEADERS = ["DeliveryID","DeliveryDate","Region","ServiceType",
           "Weight_kg","DeliveryFee","DeliveryTime_mins","CustomerRating"]
COL_WIDTHS = [12, 14, 16, 12, 12, 14, 20, 16]

def write_data_sheet(ws, include_flag_col=False):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    
    hdrs = HEADERS + (["ReviewFlag"] if include_flag_col else [])
    widths = COL_WIDTHS + ([14] if include_flag_col else [])
    
    for col_idx, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = hdr_font()
        c.fill = cell_fill(DARK)
        c.alignment = center()
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    region_colors = {
        "Nebula North":  "DDEEFF",
        "Solar South":   "E2EFDA",
        "Asteroid East": "FFF2CC",
        "Warp West":     "FCE4D6",
    }

    for row_idx, r in enumerate(records, 2):
        row_fill = cell_fill(region_colors.get(r["Region"], WHITE))
        vals = [r["DeliveryID"], r["DeliveryDate"], r["Region"], r["ServiceType"],
                r["Weight_kg"], r["DeliveryFee"], r["DeliveryTime_mins"], r["CustomerRating"]]
        if include_flag_col:
            vals.append("")  # blank for learner to fill
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = body_font()
            c.fill = row_fill
            c.border = thin_border()
            if col_idx == 2:  # date
                c.number_format = "DD/MM/YYYY"
                c.alignment = center()
            elif col_idx in [1,3,4]:
                c.alignment = left()
            else:
                c.alignment = center()
                if col_idx == 6:
                    c.number_format = '£#,##0'

# ── 1. sd2_start.xlsx ────────────────────────────────────────────────────────
print("Creating sd2_start.xlsx...")
wb = Workbook()
ws = wb.active
ws.title = "Deliveries"
ws.sheet_view.showGridLines = True
write_data_sheet(ws)

# Add a cover note at the top (shift data down) - actually just add a note sheet
ws_info = wb.create_sheet("Dataset Info")
ws_info["A1"] = "CosmoCourier Co. — Delivery Data (Jan–Jun 2025)"
ws_info["A1"].font = Font(name="Arial", size=14, bold=True, color=DARK)
ws_info["A3"] = "This dataset is cleaned and ready for analysis."
ws_info["A4"] = "72 delivery records across 4 regions and 3 service types."
ws_info["A5"] = "Date range: 1 January 2025 to 30 June 2025"
ws_info["A3"].font = body_font(11)
ws_info["A4"].font = body_font(11)
ws_info["A5"].font = body_font(11)
ws_info["A7"] = "Fields:"
ws_info["A7"].font = body_font(11, bold=True)
field_info = [
    ("DeliveryID","Unique delivery reference code"),
    ("DeliveryDate","Date the delivery was completed (DD/MM/YYYY)"),
    ("Region","Delivery region: Nebula North, Solar South, Asteroid East, Warp West"),
    ("ServiceType","Service tier: Standard, Express, or Priority"),
    ("Weight_kg","Parcel weight in kilograms"),
    ("DeliveryFee","Delivery charge in pounds (£)"),
    ("DeliveryTime_mins","Time taken to complete the delivery (minutes)"),
    ("CustomerRating","Customer satisfaction score (1=lowest, 5=highest)"),
]
for i,(fn,fd) in enumerate(field_info,8):
    ws_info[f"A{i}"] = fn
    ws_info[f"B{i}"] = fd
    ws_info[f"A{i}"].font = body_font(10, bold=True)
    ws_info[f"B{i}"].font = body_font(10)
ws_info.column_dimensions["A"].width = 22
ws_info.column_dimensions["B"].width = 60

wb.save(f"{OUT}/01_source_data/sd2_start.xlsx")
print("  ✓ sd2_start.xlsx")

# ── 2. workbook1_start.xlsx ──────────────────────────────────────────────────
print("Creating workbook1_start.xlsx...")
wb1s = Workbook()
ws_del = wb1s.active
ws_del.title = "Deliveries"
write_data_sheet(ws_del, include_flag_col=True)

# Add empty Analysis sheet with labels
ws_an = wb1s.create_sheet("Analysis")
ws_an.column_dimensions["A"].width = 32
ws_an.column_dimensions["B"].width = 20
ws_an.column_dimensions["C"].width = 30

# Title
ws_an["A1"] = "WORKBOOK 1 — ANALYSIS SHEET"
ws_an["A1"].font = Font(name="Arial", size=14, bold=True, color=DARK)
ws_an["A3"] = "PART 1: Summary Statistics"
ws_an["A3"].font = Font(name="Arial", size=12, bold=True, color=BLUE)

labels_part1 = [
    (5, "Total DeliveryFee (£)", "Enter your SUM formula here →", "B5"),
    (6, "Total Weight_kg",       "Enter your SUM formula here →", "B6"),
    (7, "Number of Deliveries",  "Enter your COUNT formula here →", "B7"),
    (8, "Average DeliveryFee (£)","Enter your AVERAGE formula here →", "B8"),
    (9, "Average CustomerRating","Enter your AVERAGE formula here →", "B9"),
    (10,"Lowest DeliveryFee (£)","Enter your MIN formula here →", "B10"),
    (11,"Highest DeliveryFee (£)","Enter your MAX formula here →", "B11"),
    (12,"Lightest Parcel (kg)",  "Enter your MIN formula here →", "B12"),
    (13,"Heaviest Parcel (kg)",  "Enter your MAX formula here →", "B13"),
]
for row,label,hint,bcell in labels_part1:
    ws_an.cell(row=row, column=1, value=label).font = body_font(10, bold=True)
    ws_an.cell(row=row, column=2, value="").fill = cell_fill(LBLUE)
    ws_an.cell(row=row, column=3, value=hint).font = Font(name="Arial",size=9,italic=True,color="888888")

ws_an["A15"] = "PART 2: Region Analysis"
ws_an["A15"].font = Font(name="Arial", size=12, bold=True, color=BLUE)

reg_labels = ["Nebula North","Solar South","Asteroid East","Warp West"]
ws_an.cell(row=17,column=1,value="Region").font = hdr_font(10)
ws_an.cell(row=17,column=1).fill = cell_fill(DARK)
ws_an.cell(row=17,column=2,value="Count (COUNTIF)").font = hdr_font(10)
ws_an.cell(row=17,column=2).fill = cell_fill(DARK)
ws_an.cell(row=17,column=3,value="Total Fee £ (SUMIF)").font = hdr_font(10)
ws_an.cell(row=17,column=3).fill = cell_fill(DARK)
ws_an.cell(row=17,column=4,value="Avg Fee £ (AVERAGEIF)").font = hdr_font(10)
ws_an.cell(row=17,column=4).fill = cell_fill(DARK)
ws_an.column_dimensions["D"].width = 22
for i,reg in enumerate(reg_labels,18):
    ws_an.cell(row=i,column=1,value=reg).font = body_font(10)
    for c in [2,3,4]:
        cell = ws_an.cell(row=i,column=c,value="")
        cell.fill = cell_fill(LBLUE)
        cell.border = thin_border()

ws_an["A23"] = "PART 3: Service Type Analysis"
ws_an["A23"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
ws_an.cell(row=25,column=1,value="ServiceType").font = hdr_font(10)
ws_an.cell(row=25,column=1).fill = cell_fill(DARK)
ws_an.cell(row=25,column=2,value="Count (COUNTIF)").font = hdr_font(10)
ws_an.cell(row=25,column=2).fill = cell_fill(DARK)
ws_an.cell(row=25,column=3,value="Total Fee £ (SUMIF)").font = hdr_font(10)
ws_an.cell(row=25,column=3).fill = cell_fill(DARK)
ws_an.cell(row=25,column=4,value="Avg Fee £ (AVERAGEIF)").font = hdr_font(10)
ws_an.cell(row=25,column=4).fill = cell_fill(DARK)
for i,svc in enumerate(["Standard","Express","Priority"],26):
    ws_an.cell(row=i,column=1,value=svc).font = body_font(10)
    for c in [2,3,4]:
        cell = ws_an.cell(row=i,column=c,value="")
        cell.fill = cell_fill(LBLUE)
        cell.border = thin_border()

ws_an["A30"] = "PART 4: Multi-Condition Analysis (COUNTIFS / SUMIFS / AVERAGEIFS)"
ws_an["A30"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
ws_an["A32"] = "Region"
ws_an["A32"].font = hdr_font(10)
ws_an["A32"].fill = cell_fill(DARK)
ws_an["B32"] = "ServiceType"
ws_an["B32"].font = hdr_font(10)
ws_an["B32"].fill = cell_fill(DARK)
ws_an["C32"] = "Count (COUNTIFS)"
ws_an["C32"].font = hdr_font(10)
ws_an["C32"].fill = cell_fill(DARK)
ws_an["D32"] = "Total Fee £ (SUMIFS)"
ws_an["D32"].font = hdr_font(10)
ws_an["D32"].fill = cell_fill(DARK)
ws_an["E32"] = "Avg Fee £ (AVERAGEIFS)"
ws_an["E32"].font = hdr_font(10)
ws_an["E32"].fill = cell_fill(DARK)
ws_an.column_dimensions["E"].width = 22
mc_rows=[("Solar South","Priority"),("Warp West","Standard"),
         ("Nebula North","Express"),("Asteroid East","Priority")]
for i,(reg,svc) in enumerate(mc_rows,33):
    ws_an.cell(row=i,column=1,value=reg).font = body_font(10)
    ws_an.cell(row=i,column=2,value=svc).font = body_font(10)
    for c in [3,4,5]:
        cell = ws_an.cell(row=i,column=c,value="")
        cell.fill = cell_fill(LBLUE)
        cell.border = thin_border()

ws_an["A38"] = "PART 5: Review Flag Count"
ws_an["A38"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
ws_an["A40"] = "Count of 'Good' records (Rating ≥ 4)"
ws_an["A40"].font = body_font(10, bold=True)
ws_an["B40"].fill = cell_fill(LBLUE)
ws_an["A41"] = "Count of 'Review' records (Rating < 4)"
ws_an["A41"].font = body_font(10, bold=True)
ws_an["B41"].fill = cell_fill(LBLUE)

wb1s.save(f"{OUT}/02_workbook_data/workbook1_start.xlsx")
print("  ✓ workbook1_start.xlsx")

# ── 3. workbook1_complete.xlsx ───────────────────────────────────────────────
print("Creating workbook1_complete.xlsx...")
wb1c = Workbook()
ws_del = wb1c.active
ws_del.title = "Deliveries"
write_data_sheet(ws_del, include_flag_col=True)

# Fill in ReviewFlag with IF formula
for row_idx in range(2, 74):
    cell = ws_del.cell(row=row_idx, column=9)
    cell.value = f'=IF(H{row_idx}<4,"Review","Good")'
    cell.font = body_font(10)
    cell.alignment = center()

ws_an = wb1c.create_sheet("Analysis")
ws_an.column_dimensions["A"].width = 32
ws_an.column_dimensions["B"].width = 20
ws_an.column_dimensions["C"].width = 30
ws_an.column_dimensions["D"].width = 22
ws_an.column_dimensions["E"].width = 22

ws_an["A1"] = "WORKBOOK 1 — ANALYSIS SHEET (COMPLETE)"
ws_an["A1"].font = Font(name="Arial", size=14, bold=True, color=DARK)

ws_an["A3"] = "PART 1: Summary Statistics"
ws_an["A3"].font = Font(name="Arial", size=12, bold=True, color=BLUE)

sum_rows = [
    (5,  "Total DeliveryFee (£)",    "=SUM(Deliveries!F2:F73)",     2653),
    (6,  "Total Weight_kg",           "=SUM(Deliveries!E2:E73)",     597.0),
    (7,  "Number of Deliveries",      "=COUNT(Deliveries!A2:A73)",   72),
    (8,  "Average DeliveryFee (£)",   "=AVERAGE(Deliveries!F2:F73)", None),
    (9,  "Average CustomerRating",    "=AVERAGE(Deliveries!H2:H73)", None),
    (10, "Lowest DeliveryFee (£)",    "=MIN(Deliveries!F2:F73)",     9),
    (11, "Highest DeliveryFee (£)",   "=MAX(Deliveries!F2:F73)",     92),
    (12, "Lightest Parcel (kg)",      "=MIN(Deliveries!E2:E73)",     2.0),
    (13, "Heaviest Parcel (kg)",      "=MAX(Deliveries!E2:E73)",     28.5),
]
for row,label,formula,_ in sum_rows:
    ws_an.cell(row=row,column=1,value=label).font = body_font(10,bold=True)
    c = ws_an.cell(row=row,column=2,value=formula)
    c.font = body_font(10,color="0000FF")
    c.fill = cell_fill(GREEN)
    c.border = thin_border()
    c.alignment = center()
    if row in [5,8,10,11]:
        c.number_format = '£#,##0.00'
    elif row in [6,12,13]:
        c.number_format = '0.0'

ws_an["A15"] = "PART 2: Region Analysis"
ws_an["A15"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
for col,label in enumerate(["Region","Count","Total Fee £","Avg Fee £"],1):
    c = ws_an.cell(row=17,column=col,value=label)
    c.font = hdr_font(10); c.fill = cell_fill(DARK)

region_data = [
    ("Nebula North",  "=COUNTIF(Deliveries!C2:C73,\"Nebula North\")",  "=SUMIF(Deliveries!C2:C73,\"Nebula North\",Deliveries!F2:F73)",  "=AVERAGEIF(Deliveries!C2:C73,\"Nebula North\",Deliveries!F2:F73)"),
    ("Solar South",   "=COUNTIF(Deliveries!C2:C73,\"Solar South\")",   "=SUMIF(Deliveries!C2:C73,\"Solar South\",Deliveries!F2:F73)",   "=AVERAGEIF(Deliveries!C2:C73,\"Solar South\",Deliveries!F2:F73)"),
    ("Asteroid East", "=COUNTIF(Deliveries!C2:C73,\"Asteroid East\")", "=SUMIF(Deliveries!C2:C73,\"Asteroid East\",Deliveries!F2:F73)", "=AVERAGEIF(Deliveries!C2:C73,\"Asteroid East\",Deliveries!F2:F73)"),
    ("Warp West",     "=COUNTIF(Deliveries!C2:C73,\"Warp West\")",     "=SUMIF(Deliveries!C2:C73,\"Warp West\",Deliveries!F2:F73)",     "=AVERAGEIF(Deliveries!C2:C73,\"Warp West\",Deliveries!F2:F73)"),
]
for i,(reg,cnt,sm,avg) in enumerate(region_data,18):
    ws_an.cell(row=i,column=1,value=reg).font=body_font(10)
    for col_idx,(val,fmt) in enumerate([(cnt,None),(sm,'£#,##0'),(avg,'£#,##0.00')],2):
        c=ws_an.cell(row=i,column=col_idx,value=val)
        c.font=body_font(10,color="0000FF"); c.fill=cell_fill(GREEN)
        c.border=thin_border(); c.alignment=center()
        if fmt: c.number_format=fmt

ws_an["A23"] = "PART 3: Service Type Analysis"
ws_an["A23"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
for col,label in enumerate(["ServiceType","Count","Total Fee £","Avg Fee £"],1):
    c=ws_an.cell(row=25,column=col,value=label)
    c.font=hdr_font(10); c.fill=cell_fill(DARK)

service_data = [
    ("Standard","=COUNTIF(Deliveries!D2:D73,\"Standard\")","=SUMIF(Deliveries!D2:D73,\"Standard\",Deliveries!F2:F73)","=AVERAGEIF(Deliveries!D2:D73,\"Standard\",Deliveries!F2:F73)"),
    ("Express", "=COUNTIF(Deliveries!D2:D73,\"Express\")", "=SUMIF(Deliveries!D2:D73,\"Express\",Deliveries!F2:F73)", "=AVERAGEIF(Deliveries!D2:D73,\"Express\",Deliveries!F2:F73)"),
    ("Priority","=COUNTIF(Deliveries!D2:D73,\"Priority\")","=SUMIF(Deliveries!D2:D73,\"Priority\",Deliveries!F2:F73)","=AVERAGEIF(Deliveries!D2:D73,\"Priority\",Deliveries!F2:F73)"),
]
for i,(svc,cnt,sm,avg) in enumerate(service_data,26):
    ws_an.cell(row=i,column=1,value=svc).font=body_font(10)
    for col_idx,(val,fmt) in enumerate([(cnt,None),(sm,'£#,##0'),(avg,'£#,##0.00')],2):
        c=ws_an.cell(row=i,column=col_idx,value=val)
        c.font=body_font(10,color="0000FF"); c.fill=cell_fill(GREEN)
        c.border=thin_border(); c.alignment=center()
        if fmt: c.number_format=fmt

ws_an["A30"] = "PART 4: Multi-Condition Analysis"
ws_an["A30"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
for col,label in enumerate(["Region","ServiceType","Count","Total Fee £","Avg Fee £"],1):
    c=ws_an.cell(row=32,column=col,value=label)
    c.font=hdr_font(10); c.fill=cell_fill(DARK)

mc_data=[
    ("Solar South","Priority",
     "=COUNTIFS(Deliveries!C2:C73,\"Solar South\",Deliveries!D2:D73,\"Priority\")",
     "=SUMIFS(Deliveries!F2:F73,Deliveries!C2:C73,\"Solar South\",Deliveries!D2:D73,\"Priority\")",
     "=AVERAGEIFS(Deliveries!F2:F73,Deliveries!C2:C73,\"Solar South\",Deliveries!D2:D73,\"Priority\")"),
    ("Warp West","Standard",
     "=COUNTIFS(Deliveries!C2:C73,\"Warp West\",Deliveries!D2:D73,\"Standard\")",
     "=SUMIFS(Deliveries!F2:F73,Deliveries!C2:C73,\"Warp West\",Deliveries!D2:D73,\"Standard\")",
     "=AVERAGEIFS(Deliveries!F2:F73,Deliveries!C2:C73,\"Warp West\",Deliveries!D2:D73,\"Standard\")"),
    ("Nebula North","Express",
     "=COUNTIFS(Deliveries!C2:C73,\"Nebula North\",Deliveries!D2:D73,\"Express\")",
     "=SUMIFS(Deliveries!F2:F73,Deliveries!C2:C73,\"Nebula North\",Deliveries!D2:D73,\"Express\")",
     "=AVERAGEIFS(Deliveries!F2:F73,Deliveries!C2:C73,\"Nebula North\",Deliveries!D2:D73,\"Express\")"),
    ("Asteroid East","Priority",
     "=COUNTIFS(Deliveries!C2:C73,\"Asteroid East\",Deliveries!D2:D73,\"Priority\")",
     "=SUMIFS(Deliveries!F2:F73,Deliveries!C2:C73,\"Asteroid East\",Deliveries!D2:D73,\"Priority\")",
     "=AVERAGEIFS(Deliveries!F2:F73,Deliveries!C2:C73,\"Asteroid East\",Deliveries!D2:D73,\"Priority\")"),
]
for i,(reg,svc,cnt,sm,avg) in enumerate(mc_data,33):
    ws_an.cell(row=i,column=1,value=reg).font=body_font(10)
    ws_an.cell(row=i,column=2,value=svc).font=body_font(10)
    for col_idx,(val,fmt) in enumerate([(cnt,None),(sm,'£#,##0'),(avg,'£#,##0.00')],3):
        c=ws_an.cell(row=i,column=col_idx,value=val)
        c.font=body_font(10,color="0000FF"); c.fill=cell_fill(GREEN)
        c.border=thin_border(); c.alignment=center()
        if fmt: c.number_format=fmt

ws_an["A38"] = "PART 5: Review Flag Count"
ws_an["A38"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
ws_an.cell(row=40,column=1,value="Count of 'Good' (Rating ≥ 4)").font=body_font(10,bold=True)
c=ws_an.cell(row=40,column=2,value='=COUNTIF(Deliveries!I2:I73,"Good")')
c.font=body_font(10,color="0000FF"); c.fill=cell_fill(GREEN); c.border=thin_border(); c.alignment=center()
ws_an.cell(row=41,column=1,value="Count of 'Review' (Rating < 4)").font=body_font(10,bold=True)
c=ws_an.cell(row=41,column=2,value='=COUNTIF(Deliveries!I2:I73,"Review")')
c.font=body_font(10,color="0000FF"); c.fill=cell_fill(GREEN); c.border=thin_border(); c.alignment=center()

wb1c.save(f"{OUT}/02_workbook_data/workbook1_complete.xlsx")
print("  ✓ workbook1_complete.xlsx")

# ── 4. workbook2_start.xlsx ──────────────────────────────────────────────────
print("Creating workbook2_start.xlsx (= WB1 complete + blank PivotWork sheet)...")
import shutil
shutil.copy(f"{OUT}/02_workbook_data/workbook1_complete.xlsx",
            f"{OUT}/02_workbook_data/workbook2_start.xlsx")
wb2s = openpyxl.load_workbook(f"{OUT}/02_workbook_data/workbook2_start.xlsx")
ws_piv = wb2s.create_sheet("PivotWork")
ws_piv["A1"] = "WORKBOOK 2 — PIVOT TABLE WORKSPACE"
ws_piv["A1"].font = Font(name="Arial", size=14, bold=True, color=DARK)
ws_piv["A3"] = "You will build your Pivot Tables, charts, slicers and timeline here."
ws_piv["A3"].font = body_font(11)
ws_piv["A4"] = "Follow workbook2_pivots_charts_slicers_timelines.docx step by step."
ws_piv["A4"].font = body_font(11)
ws_piv["A6"] = "Tip: Click the Deliveries sheet tab, then Insert → PivotTable."
ws_piv["A6"].font = Font(name="Arial", size=10, italic=True, color="555555")
ws_piv.column_dimensions["A"].width = 60
wb2s.save(f"{OUT}/02_workbook_data/workbook2_start.xlsx")
print("  ✓ workbook2_start.xlsx")

# ── 5. workbook2_complete.xlsx ───────────────────────────────────────────────
print("Creating workbook2_complete.xlsx (adds RegionSummary pivot-like table)...")
shutil.copy(f"{OUT}/02_workbook_data/workbook2_start.xlsx",
            f"{OUT}/02_workbook_data/workbook2_complete.xlsx")
wb2c = openpyxl.load_workbook(f"{OUT}/02_workbook_data/workbook2_complete.xlsx")

ws_piv = wb2c["PivotWork"]
ws_piv["A1"] = "WORKBOOK 2 — PIVOT WORK (COMPLETE)"
ws_piv["A1"].font = Font(name="Arial", size=14, bold=True, color=DARK)

ws_piv["A3"] = "PIVOT SUMMARY: Region vs Total DeliveryFee"
ws_piv["A3"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
ws_piv["A4"] = "(This table mirrors your Pivot Table. Your Pivot Table should show the same totals.)"
ws_piv["A4"].font = Font(name="Arial", size=9, italic=True, color="555555")

for col,label in enumerate(["Region","Total DeliveryFee £","Count of Deliveries","Avg DeliveryFee £"],1):
    c=ws_piv.cell(row=6,column=col,value=label)
    c.font=hdr_font(10); c.fill=cell_fill(DARK)
pivot_data=[
    ("Nebula North",618,18,34.33),
    ("Solar South",749,18,41.61),
    ("Asteroid East",716,18,39.78),
    ("Warp West",570,18,31.67),
    ("Grand Total",2653,72,36.85),
]
for i,(reg,tot,cnt,avg) in enumerate(pivot_data,7):
    bold = reg=="Grand Total"
    ws_piv.cell(row=i,column=1,value=reg).font=body_font(10,bold=bold)
    for col_idx,(val,fmt) in enumerate([(tot,'£#,##0'),(cnt,None),(avg,'£#,##0.00')],2):
        c=ws_piv.cell(row=i,column=col_idx,value=val)
        c.font=body_font(10,bold=bold,color="0000FF")
        c.fill=cell_fill(GREEN if not bold else ACCENT)
        c.border=thin_border(); c.alignment=center()
        if fmt: c.number_format=fmt

ws_piv["A13"] = "PIVOT SUMMARY: ServiceType vs Total DeliveryFee"
ws_piv["A13"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
for col,label in enumerate(["ServiceType","Total DeliveryFee £","Count","Avg Fee £"],1):
    c=ws_piv.cell(row=15,column=col,value=label)
    c.font=hdr_font(10); c.fill=cell_fill(DARK)
svc_pivot=[("Standard",340,24,14.17),("Express",872,24,36.33),("Priority",1441,24,60.04),("Grand Total",2653,72,36.85)]
for i,(svc,tot,cnt,avg) in enumerate(svc_pivot,16):
    bold=svc=="Grand Total"
    ws_piv.cell(row=i,column=1,value=svc).font=body_font(10,bold=bold)
    for col_idx,(val,fmt) in enumerate([(tot,'£#,##0'),(cnt,None),(avg,'£#,##0.00')],2):
        c=ws_piv.cell(row=i,column=col_idx,value=val)
        c.font=body_font(10,bold=bold,color="0000FF")
        c.fill=cell_fill(GREEN if not bold else ACCENT)
        c.border=thin_border(); c.alignment=center()
        if fmt: c.number_format=fmt

ws_piv["A21"] = "MONTHLY SUMMARY: DeliveryFee by Month"
ws_piv["A21"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
ws_piv["A22"] = "(Use your Timeline to filter this in your Pivot Table)"
ws_piv["A22"].font = Font(name="Arial", size=9, italic=True, color="555555")
for col,label in enumerate(["Month","Total DeliveryFee £","Count"],1):
    c=ws_piv.cell(row=24,column=col,value=label)
    c.font=hdr_font(10); c.fill=cell_fill(DARK)
month_data=[("Jan 2025",340,12),("Feb 2025",425,12),("Mar 2025",468,12),
            ("Apr 2025",433,12),("May 2025",457,12),("Jun 2025",530,12),("Total",2653,72)]
for i,(mo,tot,cnt) in enumerate(month_data,25):
    bold=mo=="Total"
    ws_piv.cell(row=i,column=1,value=mo).font=body_font(10,bold=bold)
    for col_idx,(val,fmt) in enumerate([(tot,'£#,##0'),(cnt,None)],2):
        c=ws_piv.cell(row=i,column=col_idx,value=val)
        c.font=body_font(10,bold=bold,color="0000FF")
        c.fill=cell_fill(GREEN if not bold else ACCENT)
        c.border=thin_border(); c.alignment=center()
        if fmt: c.number_format=fmt

ws_piv["A33"] = "KEY INSIGHT NOTE"
ws_piv["A33"].font = Font(name="Arial", size=12, bold=True, color=DARK)
ws_piv["A34"] = "Solar South generated the highest total delivery fees (£749)."
ws_piv["A34"].font = body_font(10)
ws_piv["A35"] = "Warp West generated the lowest total delivery fees (£570)."
ws_piv["A35"].font = body_font(10)
ws_piv["A36"] = "Priority deliveries accounted for £1,441 — more than Standard and Express combined."
ws_piv["A36"].font = body_font(10)
ws_piv["A37"] = "June 2025 was the highest-revenue month (£530). January was the lowest (£340)."
ws_piv["A37"].font = body_font(10)

for col in ["A","B","C","D"]:
    ws_piv.column_dimensions[col].width = 26

wb2c.save(f"{OUT}/02_workbook_data/workbook2_complete.xlsx")
print("  ✓ workbook2_complete.xlsx")

# ── 6. workbook3_start.xlsx ──────────────────────────────────────────────────
print("Creating workbook3_start.xlsx...")
shutil.copy(f"{OUT}/02_workbook_data/workbook2_complete.xlsx",
            f"{OUT}/02_workbook_data/workbook3_start.xlsx")
wb3s = openpyxl.load_workbook(f"{OUT}/02_workbook_data/workbook3_start.xlsx")
ws_dash = wb3s.create_sheet("Dashboard")
ws_dash["A1"] = "DASHBOARD — CosmoCourier Co."
ws_dash["A1"].font = Font(name="Arial", size=14, bold=True, color=DARK)
ws_dash["A3"] = "Build your dashboard here following workbook3_dashboard_trendline_interpretation.docx"
ws_dash["A3"].font = body_font(10, italic=True)
ws_dash["A5"] = "SUMMARY CARDS (place below the title):"
ws_dash["A5"].font = body_font(11, bold=True)
ws_dash["A6"] = "Card 1 — Total Deliveries:"
ws_dash["B6"].fill = cell_fill(LBLUE)
ws_dash["A7"] = "Card 2 — Best Region + Total:"
ws_dash["B7"].fill = cell_fill(LBLUE)
ws_dash["A8"] = "Card 3 — Overall Average Fee £:"
ws_dash["B8"].fill = cell_fill(LBLUE)
ws_scatter = wb3s.create_sheet("ScatterData")
ws_scatter["A1"] = "SCATTER CHART DATA"
ws_scatter["A1"].font = Font(name="Arial", size=12, bold=True, color=DARK)
ws_scatter["A2"] = "Weight_kg"
ws_scatter["B2"] = "DeliveryFee"
ws_scatter["A2"].font = hdr_font(10); ws_scatter["A2"].fill = cell_fill(DARK)
ws_scatter["B2"].font = hdr_font(10); ws_scatter["B2"].fill = cell_fill(DARK)
for i,r in enumerate(records,3):
    ws_scatter.cell(row=i,column=1,value=r["Weight_kg"])
    ws_scatter.cell(row=i,column=2,value=r["DeliveryFee"])
ws_scatter.column_dimensions["A"].width=14
ws_scatter.column_dimensions["B"].width=14
wb3s.save(f"{OUT}/02_workbook_data/workbook3_start.xlsx")
print("  ✓ workbook3_start.xlsx")

# ── 7. workbook3_complete.xlsx and sd2_complete.xlsx ────────────────────────
print("Creating workbook3_complete.xlsx and sd2_complete.xlsx...")
shutil.copy(f"{OUT}/02_workbook_data/workbook3_start.xlsx",
            f"{OUT}/02_workbook_data/workbook3_complete.xlsx")
wb3c = openpyxl.load_workbook(f"{OUT}/02_workbook_data/workbook3_complete.xlsx")
ws_dash = wb3c["Dashboard"]

# Complete dashboard
ws_dash["A1"] = "COSMO COURIER CO. — DELIVERY PERFORMANCE DASHBOARD"
ws_dash["A1"].font = Font(name="Arial", size=16, bold=True, color=DARK)
ws_dash["A1"].fill = cell_fill(LBLUE)

ws_dash["A3"] = "Jan–Jun 2025  |  All Regions  |  All Service Types"
ws_dash["A3"].font = body_font(10, italic=True)

# Summary cards
ws_dash["A5"] = "Total Deliveries"
ws_dash["A5"].font = Font(name="Arial", size=10, bold=True, color=WHITE)
ws_dash["A5"].fill = cell_fill(DARK)
ws_dash["A6"] = 72
ws_dash["A6"].font = Font(name="Arial", size=18, bold=True, color=DARK)
ws_dash["A6"].alignment = center()

ws_dash["C5"] = "Best Region: Solar South"
ws_dash["C5"].font = Font(name="Arial", size=10, bold=True, color=WHITE)
ws_dash["C5"].fill = cell_fill(BLUE)
ws_dash["C6"] = "£749 total fees"
ws_dash["C6"].font = Font(name="Arial", size=14, bold=True, color=BLUE)
ws_dash["C6"].alignment = center()

ws_dash["E5"] = "Overall Average Fee"
ws_dash["E5"].font = Font(name="Arial", size=10, bold=True, color=WHITE)
ws_dash["E5"].fill = cell_fill(DARK)
ws_dash["E6"] = "=PivotWork!B11"
ws_dash["E6"].font = Font(name="Arial", size=18, bold=True, color=DARK)
ws_dash["E6"].number_format = "£#,##0.00"
ws_dash["E6"].alignment = center()

ws_dash["A8"] = "Revenue by Region"
ws_dash["A8"].font = body_font(11, bold=True)
for col,label in enumerate(["Region","Total Fee £","Avg Fee £"],1):
    c=ws_dash.cell(row=9,column=col,value=label)
    c.font=hdr_font(10); c.fill=cell_fill(DARK)
dash_reg=[("Nebula North",618,34.33),("Solar South",749,41.61),
          ("Asteroid East",716,39.78),("Warp West",570,31.67)]
for i,(reg,tot,avg) in enumerate(dash_reg,10):
    ws_dash.cell(row=i,column=1,value=reg).font=body_font(10)
    for col_idx,(val,fmt) in enumerate([(tot,'£#,##0'),(avg,'£#,##0.00')],2):
        c=ws_dash.cell(row=i,column=col_idx,value=val)
        c.font=body_font(10,color="0000FF"); c.fill=cell_fill(GREEN)
        c.border=thin_border(); c.alignment=center()
        if fmt: c.number_format=fmt

ws_dash["A15"] = "Trendline: Weight vs DeliveryFee"
ws_dash["A15"].font = body_font(11, bold=True)
ws_dash["A16"] = "Equation: y = 2.80x + 13.64"
ws_dash["A16"].font = body_font(10)
ws_dash["A17"] = "R² = 0.34 — Moderate fit. Weight explains ~34% of fee variation."
ws_dash["A17"].font = body_font(10)

ws_dash["A19"] = "INTERPRETATION"
ws_dash["A19"].font = Font(name="Arial", size=12, bold=True, color=DARK)
ws_dash["A20"] = "The data shows that Solar South is the highest-revenue region (£749), while Warp West"
ws_dash["A21"] = "performs weakest (£570) and also has the lowest customer satisfaction scores."
ws_dash["A22"] = "Priority service generates the most revenue (£1,441 — 54% of total fees)."
ws_dash["A23"] = "Revenue grew overall from January (£340) to June (£530), suggesting healthy demand growth."
ws_dash["A24"] = "The trendline shows heavier parcels generate higher fees (y = 2.80x + 13.64), but the R²"
ws_dash["A25"] = "of 0.34 means service type also plays a major role in determining the final fee."
for row in range(20,26):
    ws_dash.cell(row=row,column=1).font=body_font(10)

for col in ["A","B","C","D","E","F"]:
    ws_dash.column_dimensions[col].width = 22

wb3c.save(f"{OUT}/02_workbook_data/workbook3_complete.xlsx")
print("  ✓ workbook3_complete.xlsx")
shutil.copy(f"{OUT}/02_workbook_data/workbook3_complete.xlsx",
            f"{OUT}/02_workbook_data/sd2_complete.xlsx")
print("  ✓ sd2_complete.xlsx")
print("\nAll Excel files created successfully!")
